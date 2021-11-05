from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import *
import time


def exel(times, result, path_to_file, list_name):
    print('write_to_exel started work')

    try:
        wb = load_workbook(path_to_file)
        try:
            work_sheet = wb[list_name]
        except:
            work_sheet = wb.create_sheet(list_name)
    except:
        wb = Workbook()
        work_sheet = wb.create_sheet(list_name)

    last_column = 1

# Find out the last empty column
    while True:
        new_column = work_sheet.cell(row=1, column=last_column)
        if new_column.value != None:
            last_column = last_column + 1
        elif new_column.value == None and last_column == 1:
            work_cell_for_case = work_sheet.cell(row=1, column=1)
            work_cell_for_case.value = 'Name'
            last_column = 2
            break
        else:
            break
    # print(last_column)


    # Adding a date and time cell
    work_cell_for_case = work_sheet.cell(row=1, column=last_column)
    work_cell_for_case.value = times
    work_cell_for_time = work_sheet.cell(row=1, column=last_column + 1)
    work_cell_for_time.value = "Время"


# Looks at who is already recorded in the table and adds the received data about the completed cases
    i = 2
    empty = 5
    while True:
        name = work_sheet.cell(row=i, column=1).value
        if name in result.keys():
            work_cell_for_case = work_sheet.cell(row=i, column=last_column)
            work_cell_for_case.value = result[name][0]
            work_cell_for_time = work_sheet.cell(row=i, column=last_column + 1)
            if result[name][1] < 86400:
                work_cell_for_time.value = time.strftime("%H:%M", time.gmtime(result[name][1]))
            else:
                work_cell_for_time.value = time.strftime("%d:%H:%M", time.gmtime(result[name][1] - 86400))
            del result[name]
            print(result)
            empty = 5
            # print(name)
        elif name not in result.keys() and name != None:
            work_cell_for_case = work_sheet.cell(row=i, column=last_column)
            work_cell_for_case.value = 0
            empty = 5
            # print(name)
        elif name == None:
            if empty != 0:
                empty = empty - 1
            else:
                break
        i = i + 1

# If suddenly there is someone who is not in the plate-adds his name 10 lines from the rest
    result_local = result
    if len(result_local) > 0:
        for name in result_local:
            work_cell_for_case = work_sheet.cell(row=i, column=1)
            work_cell_for_case.value = name
            work_cell_for_case = work_sheet.cell(row=i, column=last_column)
            work_cell_for_case.value = result[name][0]
            work_cell_for_time = work_sheet.cell(row=i, column=last_column + 1)
            if result[name][1] < 86400:
                work_cell_for_time.value = time.strftime("%H:%M", time.gmtime(result[name][1]))
            else:
                work_cell_for_time.value = time.strftime("%d:%H:%M", time.gmtime(result[name][1] - 86400))
            # del result[name]
            i = i + 1
            # print("Я тут 5")

# Saves the file with the changes made
    print('write_to_exel writes value to excel')
    try:
        wb.save(path_to_file)
    except:
        print('----------------------------------------------------------------')
        print("Тебе говорили держи Excel закрытым, во время работы программы?! "
              "Все херня, давай по новой. И закрыть Excel не забудь!")
        print('----------------------------------------------------------------')
        exit()
print('write_to_exel finishes work successfully')





def debug():
    time = datetime.now().strftime('%d.%m.%y %H:%M')
    result = {'Anastasia Vlasova': [150, 100200, 80414], 'Anatoly Petrov': [214, 230400, 205583], 'Andrey Lysenkov': [291, 364500, 277651]}
    path_to_file = ("C:/Users/ivan.shakirov/Desktop/Excel.xlsx")
    list_name = 'FRS'
    exel(time, result, path_to_file, list_name)

# debug()
